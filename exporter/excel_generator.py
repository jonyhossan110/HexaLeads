import os
from datetime import date
from pathlib import Path
from typing import Dict, Iterable, List, Optional

from openpyxl import Workbook

DEFAULT_COLUMNS = [
    "Name",
    "Address",
    "Phone",
    "Website",
    "Website Status",
    "Email",
    "Facebook",
    "LinkedIn",
    "Score",
    "Recommendation",
]


class ExcelGenerator:
    def __init__(self, download_folder: Optional[Path] = None):
        self.download_folder = Path(download_folder) if download_folder else self._default_download_folder()
        self.download_folder.mkdir(parents=True, exist_ok=True)

    @staticmethod
    def _default_download_folder() -> Path:
        home = Path.home()
        return home / "Downloads" / "HexaLeads"

    @staticmethod
    def _safe_string(value: Optional[object]) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def save(self, country: str, city: str, category: str, leads: List[Dict[str, object]], output_folder: Optional[Path] = None) -> Path:
        destination_folder = Path(output_folder) if output_folder else self.download_folder
        destination_folder.mkdir(parents=True, exist_ok=True)
        file_name = self._build_filename(country, city, category)
        output_path = destination_folder / file_name

        workbook = Workbook()
        self._write_sheet(workbook.active, "All Leads", leads)
        workbook.active.title = "All Leads"
        self._write_sheet(workbook.create_sheet("Hot Leads"), [lead for lead in leads if self._is_hot_lead(lead)])
        self._write_sheet(workbook.create_sheet("Website Needed"), [lead for lead in leads if not self._safe_string(lead.get("Website", lead.get("website"))).strip()])
        self._write_metadata_sheet(workbook, country, city, category)
        workbook.save(output_path)
        return output_path

    def generate_summary(self, leads: Iterable[Dict[str, object]]) -> Dict[str, object]:
        lead_list = list(leads)
        total = len(lead_list)
        hot = sum(1 for lead in lead_list if self._is_hot_lead(lead))
        website_missing = sum(1 for lead in lead_list if not self._safe_string(lead.get("Website", lead.get("website"))).strip())
        return {
            "total_leads": total,
            "hot_leads": hot,
            "website_needed": website_missing,
        }

    def _write_sheet(self, sheet, leads: List[Dict[str, object]]) -> None:
        sheet.append(DEFAULT_COLUMNS)
        for lead in leads:
            sheet.append([
                self._safe_string(lead.get("Name", lead.get("name"))),
                self._safe_string(lead.get("Address", lead.get("address"))),
                self._safe_string(lead.get("Phone", lead.get("phone"))),
                self._safe_string(lead.get("Website", lead.get("website"))),
                self._safe_string(lead.get("Website Status", lead.get("website_status"))),
                self._safe_string(lead.get("Email", lead.get("email"))),
                self._safe_string(lead.get("Facebook", lead.get("facebook"))),
                self._safe_string(lead.get("LinkedIn", lead.get("linkedin"))),
                self._safe_string(lead.get("Score", lead.get("score"))),
                self._safe_string(lead.get("Recommendation", lead.get("recommendation"))),
            ])

    def _write_metadata_sheet(self, workbook, country: str, city: str, category: str) -> None:
        metadata = workbook.create_sheet("Metadata")
        metadata.append(["Report Title", "HexaLeads Report — HexaCyberLab"])
        metadata.append(["Creator", "Md. Jony Hassain"])
        metadata.append(["Agency", "HexaCyberLab"])
        metadata.append(["LinkedIn", "https://www.linkedin.com/in/md-jony-hassain"])
        metadata.append(["Generated On", date.today().isoformat()])
        metadata.append(["Report by", "Md. Jony Hassain | LinkedIn: md-jony-hassain"])
        metadata.append(["Country", country])
        metadata.append(["City", city])
        metadata.append(["Category", category])

    @staticmethod
    def _build_filename(country: str, city: str, category: str) -> str:
        safe_country = ExcelGenerator._sanitize_filename_part(country)
        safe_city = ExcelGenerator._sanitize_filename_part(city)
        safe_category = ExcelGenerator._sanitize_filename_part(category)
        date_string = date.today().isoformat()
        return f"{safe_country}_{safe_city}_{safe_category}_{date_string}.xlsx"

    @staticmethod
    def _sanitize_filename_part(value: str) -> str:
        return "".join(char if char.isalnum() or char in "-_" else "_" for char in str(value).strip()) or "unknown"

    @staticmethod
    def _is_hot_lead(lead: Dict[str, object]) -> bool:
        score = lead.get("Score", lead.get("score", 0))
        try:
            return float(score) > 70
        except (TypeError, ValueError):
            return False


def build_telegram_notification(report_path: Path, total_leads: int, hot_leads: int) -> str:
    return (
        "✅ Hunting complete!\n"
        f"Found: {total_leads} businesses\n"
        f"🔥 Hot Leads: {hot_leads}\n"
        f"📁 File saved to: {report_path}\n"
        "Download: /download command"
    )
