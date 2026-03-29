import asyncio
import json
import shutil
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

from openpyxl import Workbook

from brain.brain_engine import deep_search_contact, is_non_aggregator_url, verify_lead_text
from brain.console_ui import log_brain_thought, print_mission_table
from brain.lead_analyzer import LeadAnalyzer
from brain.website_checker import WebsiteChecker

from hunter.multi_search import MultiSearchScraper

from planner import (
    MISSION_STEPS,
    format_step_completed,
    format_step_in_progress,
    init_task_tracker,
    load_task_tracker,
    mark_step_completed,
    resume_from_step,
)

SAFE_PATH_RE = __import__("re").compile(r"[^a-zA-Z0-9_-]")


@dataclass
class ScraperJob:
    country: str
    city: str
    category: str
    filters: Dict[str, Any] = field(default_factory=dict)
    mission_plan: Dict[str, Any] = field(default_factory=dict)
    stage: str = "queued"
    progress: int = 0
    message: str = "Queued"
    started_at: Optional[str] = None
    updated_at: Optional[str] = None

    def as_dict(self) -> Dict[str, Any]:
        return {
            "country": self.country,
            "city": self.city,
            "category": self.category,
            "filters": self.filters,
            "mission_plan": self.mission_plan,
            "stage": self.stage,
            "progress": self.progress,
            "message": self.message,
            "started_at": self.started_at,
            "updated_at": self.updated_at,
        }


class ScraperManager:
    def __init__(
        self,
        base_dir: Path,
        logger: Optional[Callable[[str], None]] = None,
        status_callback: Optional[Callable[[Dict[str, Any]], None]] = None,
        completion_callback: Optional[Callable[[Dict[str, Any]], None]] = None,
    ):
        self.base_dir = base_dir
        self.data_dir = self.base_dir / "data"
        self.pipeline_script = self.base_dir / "src" / "pipeline" / "pipeline.py"
        self.job_state_file = "job_status.json"
        self.output_dir = self.base_dir / "output"
        self.logger = logger
        self.status_callback = status_callback
        self.completion_callback = completion_callback
        self.queue: asyncio.Queue[ScraperJob] = asyncio.Queue()
        self.active_job: Optional[ScraperJob] = None
        self.worker_task: Optional[asyncio.Task] = None
        self._website_checker = WebsiteChecker()

    def _safe(self, value: str) -> str:
        return SAFE_PATH_RE.sub("_", value.strip()) or "unknown"

    def project_folder(self, country: str, city: str, category: str) -> Path:
        folder = self.data_dir / self._safe(country) / self._safe(city) / self._safe(category)
        folder.mkdir(parents=True, exist_ok=True)
        return folder

    def _job_status_path(self, folder: Path) -> Path:
        return folder / self.job_state_file

    @staticmethod
    def _coerce_status_text(msg: Any) -> str:
        if msg is None:
            return "No details."
        if isinstance(msg, bool):
            return str(msg)
        s = str(msg).strip()
        return s if s else "No details."

    def _log(self, message: str) -> None:
        if self.logger:
            self.logger(message)

    def _rich_error(self, message: str, exc: BaseException) -> None:
        try:
            from rich.console import Console

            Console().print(f"[bold red]{message}[/bold red]: {exc}")
        except Exception:
            print(f"{message}: {exc}")

    def _notify_completion(
        self,
        job: ScraperJob,
        folder: Path,
        *,
        success: bool,
        error: Optional[str] = None,
    ) -> None:
        if not self.completion_callback:
            return
        tagged = self._load_json(folder / "final_leads.json") or []
        n = len(tagged) if isinstance(tagged, list) else 0
        xlsx = folder / "report.xlsx"
        pdf = folder / "report.pdf"
        payload: Dict[str, Any] = {
            "success": success,
            "country": job.country,
            "city": job.city,
            "category": job.category,
            "lead_count": n,
            "report_xlsx": str(xlsx) if xlsx.exists() else None,
            "report_pdf": str(pdf) if pdf.exists() else None,
            "error": error,
        }
        try:
            self.completion_callback(payload)
        except Exception as exc:
            self._rich_error("completion_callback", exc)

    def _update_status(
        self,
        job: ScraperJob,
        stage: str,
        progress: int,
        message: str,
        *,
        step_id: Optional[int] = None,
        kind: str = "info",
    ) -> None:
        message = self._coerce_status_text(message)
        job.stage = stage
        job.progress = progress
        job.message = message
        job.updated_at = datetime.utcnow().isoformat() + "Z"
        if job.started_at is None:
            job.started_at = job.updated_at
        status = job.as_dict()
        status["step_id"] = step_id
        status["kind"] = kind
        if self.status_callback:
            self.status_callback(status)
        self._save_json(self._job_status_path(self.project_folder(job.country, job.city, job.category)), status)
        self._log(f"STATUS: {stage.upper()} - {message}")

    def _save_json(self, path: Path, payload: Any) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8") as handle:
            json.dump(payload, handle, indent=2)

    def _load_json(self, path: Path) -> Any:
        if not path.exists():
            return None
        try:
            with path.open("r", encoding="utf-8") as handle:
                return json.load(handle)
        except Exception:
            return None

    def load_job_status(self, country: str, city: str, category: str) -> Optional[Dict[str, Any]]:
        folder = self.project_folder(country, city, category)
        return self._load_json(self._job_status_path(folder))

    async def enqueue_job(
        self,
        country: str,
        city: str,
        category: str,
        filters: Dict[str, Any],
        mission_plan: Optional[Dict[str, Any]] = None,
    ) -> ScraperJob:
        job = ScraperJob(
            country=country,
            city=city,
            category=category,
            filters=filters,
            mission_plan=dict(mission_plan or {}),
        )
        folder = self.project_folder(country, city, category)
        existing = load_task_tracker(folder)
        if existing and int(existing.get("last_completed_step", 0)) >= 6:
            path = folder / "task_tracker.json"
            try:
                path.unlink(missing_ok=True)
            except OSError:
                pass

        await self.queue.put(job)
        self._log(f"Queued job: {country}/{city}/{category}")
        return job

    async def run(self) -> None:
        if self.worker_task is not None:
            return
        self.worker_task = asyncio.create_task(self._worker())

    async def _worker(self) -> None:
        while True:
            job = await self.queue.get()
            self.active_job = job
            try:
                await self._process_job(job)
            except Exception as exc:
                self._rich_error("Mission failed (unexpected)", exc)
                self._log(f"ERROR: {exc}")
            finally:
                self.active_job = None
                self.queue.task_done()

    async def _process_job(self, job: ScraperJob) -> None:
        folder = self.project_folder(job.country, job.city, job.category)
        completion_error: Optional[str] = None
        completion_ok = True
        try:
            self._ensure_project_files(folder)

            mission_dict = job.mission_plan or {
                "intent": "hunt_leads",
                "category": job.category,
                "city": job.city,
                "country": job.country,
            }
            if load_task_tracker(folder) is None:
                init_task_tracker(folder, mission_dict)

            try:
                rows = [
                    (str(s["id"]), s["title"].split("—")[0].strip(), "pending")
                    for s in MISSION_STEPS
                ]
                print_mission_table(f"{job.category} · {job.city}, {job.country}", rows)
            except Exception:
                pass

            next_step = resume_from_step(folder)

            if next_step <= 1:
                await self._mission_step_1_validate(job, folder)
            if next_step <= 2:
                await self._mission_step_2_scrape(job, folder)
            if next_step <= 3:
                await self._mission_step_3_enrich(job, folder)
            if next_step <= 4:
                try:
                    await self._mission_step_4_score(job, folder)
                except Exception as exc:
                    self._rich_error("Step 4 scoring", exc)
                    osint_fallback = self._load_json(folder / "osint.json") or []
                    if isinstance(osint_fallback, list):
                        self._save_json(folder / "final_leads.json", osint_fallback)
                    mark_step_completed(folder, 4, f"Step 4 error (continuing): {exc}")
            if next_step <= 5:
                try:
                    await self._mission_step_5_local_brain(job, folder)
                except Exception as exc:
                    self._rich_error("Step 5 Local Brain", exc)
                    mark_step_completed(folder, 5, f"Step 5 error (continuing): {exc}")
            if next_step <= 6:
                try:
                    await self._mission_step_6_reports(job, folder)
                except Exception as exc:
                    self._rich_error("Step 6 reports", exc)
                    mark_step_completed(folder, 6, f"Step 6 error: {exc}")

            self._update_status(job, "completed", 100, "✅ Job completed successfully", step_id=6)
        except Exception as exc:
            completion_error = str(exc)
            completion_ok = False
            self._rich_error("Mission pipeline", exc)
            try:
                self._update_status(
                    job,
                    "failed",
                    0,
                    f"⚠️ Mission error: {exc}",
                    kind="error",
                )
            except Exception:
                pass
        finally:
            self._notify_completion(job, folder, success=completion_ok, error=completion_error)

    def _ensure_project_files(self, folder: Path) -> None:
        for file_name in ["leads.json", "analyzed.json", "osint.json", "final_leads.json"]:
            path = folder / file_name
            if not path.exists():
                self._save_json(path, [])

    def _load_leads(self, folder: Path) -> List[Dict[str, Any]]:
        leads = self._load_json(folder / "leads.json")
        return leads if isinstance(leads, list) else []

    def _save_leads(self, folder: Path, leads: List[Dict[str, Any]]) -> None:
        self._save_json(folder / "leads.json", leads)

    def _scrape_limit(self, job: ScraperJob) -> int:
        lim = job.filters.get("limit", 10)
        try:
            return max(1, int(lim))
        except (TypeError, ValueError):
            return 10

    def _print_step_done(self, _job: ScraperJob, step_id: int, detail: str) -> None:
        line = format_step_completed(step_id, detail)
        self._log(line)
        log_brain_thought(line, style="green")
        progress = min(round(100 * step_id / 6), 100)
        if self.status_callback:
            self.status_callback(
                {
                    "stage": f"step_{step_id}",
                    "progress": progress,
                    "message": line,
                    "step_id": step_id,
                    "kind": "step_done",
                }
            )

    def _copy_output_file(self, filename: str, folder: Path, dest_name: Optional[str] = None) -> Path:
        src = self.output_dir / filename
        if not src.exists():
            raise FileNotFoundError(f"Expected pipeline output missing: {src}")
        dest = folder / (dest_name or filename)
        dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(src, dest)
        return dest

    async def _run_pipeline_phase(self, job: ScraperJob, phase: str, *, step_id: int) -> None:
        if not self.pipeline_script.exists():
            raise FileNotFoundError("Pipeline script is missing.")
        limit = self._scrape_limit(job)
        command = [
            sys.executable,
            str(self.pipeline_script),
            "--city",
            job.city,
            "--keyword",
            job.category,
            "--limit",
            str(limit),
            "--phase",
            phase,
        ]
        pct = min(15 + {"scrape": 0, "analyze": 15, "osint": 30, "score": 50}.get(phase, 0), 85)
        last_error: Optional[RuntimeError] = None
        for attempt in (1, 2):
            log_brain_thought(
                f"Pipeline [{phase}] attempt {attempt}/2 — subprocess: {' '.join(command[-4:])}",
                style="yellow" if attempt > 1 else "cyan",
            )
            self._update_status(
                job,
                phase,
                pct,
                format_step_in_progress(step_id, f"Pipeline: {phase} (attempt {attempt})"),
                step_id=step_id,
            )
            # asyncio loop.subprocess_exec rejects text=True (raises ValueError: text must be False).
            process = await asyncio.create_subprocess_exec(
                *command,
                cwd=str(self.base_dir),
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.STDOUT,
            )
            while True:
                line = await process.stdout.readline()
                if not line:
                    break
                try:
                    decoded = line.decode("utf-8", errors="replace").rstrip()
                except Exception:
                    decoded = str(line).rstrip()
                if decoded:
                    self._log(decoded)
            code = await process.wait()
            if code == 0:
                return
            last_error = RuntimeError(f"Pipeline phase {phase} failed with exit code {code}")
            log_brain_thought(f"Self-fix: retrying phase {phase} after failure.", style="red")
        assert last_error is not None
        raise last_error

    async def _mission_step_1_validate(self, job: ScraperJob, folder: Path) -> None:
        log_brain_thought(format_step_in_progress(1, "Checking mission parameters."), style="magenta")
        self._update_status(job, "validate", 5, format_step_in_progress(1), step_id=1)
        if not job.category.strip() or not job.city.strip() or not job.country.strip():
            raise ValueError("Invalid mission: category, city, and country are required.")
        mark_step_completed(folder, 1, "Target validated")
        self._print_step_done(job, 1, f"Target OK — {job.category} in {job.city}, {job.country}")

    def _hybrid_enrich_leads(self, job: ScraperJob, folder: Path) -> None:
        """Maps names + hybrid search (DuckDuckGo/Bing) for sites and LinkedIn."""
        leads = self._load_leads(folder)
        if not leads:
            return
        scraper = MultiSearchScraper(delay=1.2, timeout=14)
        city, country, cat = job.city, job.country, job.category
        out: List[Dict[str, Any]] = []
        for idx, lead in enumerate(leads):
            if not isinstance(lead, dict):
                continue
            try:
                name = str(lead.get("name") or "").strip()
                merged = dict(lead)
                if not name:
                    out.append(merged)
                    continue
                query = f'"{name}" {city} {country} {cat} official website OR linkedin'
                results = scraper.search_single_query(query)
                for r in results[:18]:
                    url = getattr(r, "url", "") or ""
                    if not url:
                        continue
                    low = url.lower()
                    if "linkedin.com" in low and not merged.get("linkedin"):
                        merged["linkedin"] = url.split("?")[0].rstrip("/")
                    elif not merged.get("website") and is_non_aggregator_url(url):
                        merged["website"] = url
                out.append(merged)
            except Exception as exc:
                self._rich_error(f"Hybrid enrich lead {idx}", exc)
                out.append(lead if isinstance(lead, dict) else {})
        self._save_leads(folder, out)

    async def _mission_step_2_scrape(self, job: ScraperJob, folder: Path) -> None:
        self._update_status(job, "scraping", 15, format_step_in_progress(2, "Maps + aggregators (visible browser)."), step_id=2)
        try:
            await self._run_pipeline_phase(job, "scrape", step_id=2)
            self._copy_output_file("businesses.json", folder, "leads.json")
            try:
                self._hybrid_enrich_leads(job, folder)
            except Exception as exc:
                self._rich_error("Hybrid enrichment (post-maps)", exc)
            raw = self._load_json(folder / "leads.json") or []
            count = len(raw) if isinstance(raw, list) else 0
            mark_step_completed(folder, 2, f"{count} raw leads merged")
            self._print_step_done(job, 2, f"{count} leads found (merged)")
        except Exception as exc:
            self._rich_error("Step 2 scrape", exc)
            self._save_json(folder / "leads.json", [])
            mark_step_completed(folder, 2, f"Step 2 error (continuing): {exc}")
            self._print_step_done(job, 2, "0 leads (step failed, continuing)")

    async def _mission_step_3_enrich(self, job: ScraperJob, folder: Path) -> None:
        self._update_status(job, "enrich", 40, format_step_in_progress(3, "Go analyzer + Python OSINT."), step_id=3)
        try:
            await self._run_pipeline_phase(job, "analyze", step_id=3)
            self._copy_output_file("analyzed.json", folder)
        except Exception as exc:
            self._rich_error("Step 3 analyze", exc)
            leads = self._load_leads(folder)
            self._save_json(folder / "analyzed.json", leads)
        try:
            await self._run_pipeline_phase(job, "osint", step_id=3)
            self._copy_output_file("osint.json", folder)
        except Exception as exc:
            self._rich_error("Step 3 osint", exc)
            analyzed = self._load_json(folder / "analyzed.json") or []
            self._save_json(folder / "osint.json", analyzed if isinstance(analyzed, list) else [])
        osint_list = self._load_json(folder / "osint.json") or []
        n = len(osint_list) if isinstance(osint_list, list) else 0
        mark_step_completed(folder, 3, f"Enriched {n} leads")
        self._print_step_done(job, 3, f"Deep enrichment done for {n} leads")

    async def _mission_step_4_score(self, job: ScraperJob, folder: Path) -> None:
        self._update_status(job, "scoring", 60, format_step_in_progress(4, "Scoring engine."), step_id=4)
        await self._run_pipeline_phase(job, "score", step_id=4)
        self._copy_output_file("final_leads.json", folder)
        scored = self._load_json(folder / "final_leads.json") or []
        n = len(scored) if isinstance(scored, list) else 0
        mark_step_completed(folder, 4, f"Scored {n} leads")
        self._print_step_done(job, 4, f"Scoring complete ({n} leads)")

    def _merge_scores_into_osint(self, osint_leads: List[Dict[str, Any]], scored: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        by_site: Dict[str, Dict[str, Any]] = {}
        for row in scored:
            if not isinstance(row, dict):
                continue
            key = str(row.get("website", "") or row.get("name", "")).strip().lower()
            if key:
                by_site[key] = row
        merged: List[Dict[str, Any]] = []
        for lead in osint_leads:
            if not isinstance(lead, dict):
                continue
            copy = dict(lead)
            k = str(copy.get("website", "") or copy.get("name", "")).strip().lower()
            if k and k in by_site:
                srow = by_site[k]
                if srow.get("score") is not None:
                    copy["score"] = srow.get("score")
                if srow.get("priority"):
                    copy["priority"] = srow.get("priority")
            merged.append(copy)
        return merged

    def _apply_local_brain(
        self,
        job: ScraperJob,
        lead: Dict[str, Any],
    ) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
        name = str(lead.get("name") or "").strip()
        website = str(lead.get("website") or "").strip()
        if not name and not website:
            return None, "Missing name and website (junk)."

        enriched = dict(lead)
        if not enriched.get("email") or not enriched.get("website"):
            try:
                extra = deep_search_contact(enriched, job.city, job.country, job.category)
                for k, v in extra.items():
                    if not v:
                        continue
                    if k in ("email", "website", "linkedin") and not enriched.get(k):
                        enriched[k] = v
                    elif str(k).startswith("deep_"):
                        enriched[k] = v
            except Exception as exc:
                self._rich_error("Deep search (Local Brain)", exc)

        ok, nlp_reason = verify_lead_text(enriched, job.category)
        if not ok:
            return None, nlp_reason
        enriched["spacy_gate"] = nlp_reason

        if website:
            chk = self._website_checker.check_url(website)
            enriched["website_status"] = chk.get("status")
            if chk.get("status") == "dead":
                return None, f"Broken website: {chk.get('reason', 'unreachable')}"
        analyzer = LeadAnalyzer(job.category)
        analysis = analyzer.analyze(enriched)
        reasons = analysis.get("reasons") or []
        blob = " ".join(str(r) for r in reasons).lower()
        if "larger company" in blob or "corporation" in blob:
            return None, "Filtered as non-target (junk/large company)."
        enriched["brain_label"] = analysis.get("label")
        enriched["brain_reasons"] = reasons
        if analysis.get("score") is not None:
            enriched["score"] = analysis.get("score")
        return enriched, None

    async def _mission_step_5_local_brain(self, job: ScraperJob, folder: Path) -> None:
        self._update_status(job, "local_brain", 75, format_step_in_progress(5, "spaCy + LeadAnalyzer."), step_id=5)
        osint_leads = self._load_json(folder / "osint.json") or []
        scored = self._load_json(folder / "final_leads.json") or []
        if not isinstance(osint_leads, list):
            osint_leads = []
        if not isinstance(scored, list):
            scored = []

        merged = self._merge_scores_into_osint(osint_leads, scored)
        kept: List[Dict[str, Any]] = []
        skipped: List[Dict[str, Any]] = []
        for lead in merged:
            out, reason = self._apply_local_brain(job, lead)
            if out is None:
                skipped.append(
                    {
                        "name": lead.get("name", ""),
                        "website": lead.get("website", ""),
                        "reason": reason,
                    }
                )
                continue
            kept.append(out)

        kept = self.filter_leads(kept, job.filters)
        tagged = [self.tag_lead(lead) for lead in kept]
        self._save_json(folder / "final_leads.json", tagged)
        self._save_json(folder / "skip_log.json", skipped)
        mark_step_completed(folder, 5, f"Filtered to {len(tagged)} leads ({len(skipped)} skipped)")
        self._print_step_done(job, 5, f"{len(tagged)} leads kept ({len(skipped)} skipped by Local Brain + spaCy)")

    async def _mission_step_6_reports(self, job: ScraperJob, folder: Path) -> None:
        self._update_status(job, "finalizing", 88, format_step_in_progress(6, "Excel + PDF."), step_id=6)
        tagged = self._load_json(folder / "final_leads.json") or []
        if not isinstance(tagged, list):
            tagged = []
        self.generate_excel_report(folder / "report.xlsx", tagged)
        self.generate_pdf_report(folder / "report.pdf", tagged)
        mark_step_completed(folder, 6, "Reports written")
        self._print_step_done(job, 6, f"Excel & PDF saved ({len(tagged)} rows)")
        self._update_status(job, "finalizing", 95, "Final reports generated", step_id=6)

    def filter_leads(self, leads: List[Dict[str, Any]], filters: Dict[str, Any]) -> List[Dict[str, Any]]:
        filtered = []
        for lead in leads:
            if filters.get("has_website") is True and not lead.get("website"):
                continue
            if filters.get("has_email") is True and not lead.get("email"):
                continue
            if filters.get("has_linkedin") is True and not lead.get("linkedin"):
                continue
            score_threshold = filters.get("score_threshold")
            if score_threshold is not None:
                try:
                    score_value = float(lead.get("score", 0) or 0)
                except (TypeError, ValueError):
                    score_value = 0
                if score_value < float(score_threshold):
                    continue
            filtered.append(lead)
        return filtered

    def tag_lead(self, lead: Dict[str, Any]) -> Dict[str, Any]:
        lead = lead.copy()
        score = float(lead.get("score", 0) or 0)
        email = bool(lead.get("email"))
        website = bool(lead.get("website"))

        if score >= 80 and email:
            lead["priority"] = "HOT"
        elif score >= 50 or (email and website):
            lead["priority"] = "WARM"
        else:
            lead["priority"] = "COLD"
        return lead

    def generate_excel_report(self, path: Path, leads: List[Dict[str, Any]]) -> None:
        workbook = Workbook()
        all_sheet = workbook.active
        all_sheet.title = "All Leads"
        self._write_sheet(all_sheet, leads, [
            "Name",
            "Category",
            "Website",
            "Email Status",
            "Phone",
            "Social Links",
            "Score",
            "Priority",
        ])

        hot_sheet = workbook.create_sheet("High Priority")
        high_leads = [lead for lead in leads if lead.get("priority") == "HOT"]
        self._write_sheet(hot_sheet, high_leads, [
            "Name",
            "Category",
            "Website",
            "Email Status",
            "Phone",
            "Social Links",
            "Score",
            "Priority",
        ])

        contact_sheet = workbook.create_sheet("Contact Info")
        self._write_sheet(contact_sheet, leads, [
            "Name",
            "Email",
            "Phone",
            "Website",
            "LinkedIn",
            "Facebook",
            "X",
            "Priority",
        ], contact_mode=True)

        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(str(path))

    def generate_pdf_report(self, path: Path, leads: List[Dict[str, Any]]) -> None:
        try:
            from fpdf import FPDF
        except ImportError:
            self._log("fpdf2 is not installed; skipping PDF export.")
            return
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font("helvetica", "B", 14)
        pdf.cell(0, 10, "HexaLeads Report", ln=1)
        pdf.set_font("helvetica", size=9)
        pdf.cell(0, 6, f"Generated: {datetime.utcnow().isoformat()}Z", ln=1)
        pdf.ln(4)
        pdf.set_font("helvetica", size=8)
        for lead in leads[:500]:
            name = str(lead.get("name", ""))[:90]
            website = str(lead.get("website", ""))[:90]
            score = lead.get("score", "")
            line = f"{name} | {website} | score={score}"
            safe = line.encode("ascii", "replace").decode("ascii")
            pdf.multi_cell(0, 5, safe)
        path.parent.mkdir(parents=True, exist_ok=True)
        pdf.output(str(path))

    def _write_sheet(self, sheet: Any, leads: List[Dict[str, Any]], headers: List[str], contact_mode: bool = False) -> None:
        sheet.append(headers)
        for lead in leads:
            if contact_mode:
                sheet.append([
                    lead.get("name", ""),
                    lead.get("email", ""),
                    lead.get("phone", ""),
                    lead.get("website", ""),
                    lead.get("linkedin", ""),
                    lead.get("facebook", ""),
                    lead.get("x", ""),
                    lead.get("priority", ""),
                ])
            else:
                social_links = ", ".join(
                    filter(None, [lead.get("linkedin"), lead.get("facebook"), lead.get("x")])
                )
                sheet.append([
                    lead.get("name", ""),
                    lead.get("category", ""),
                    lead.get("website", ""),
                    "Yes" if lead.get("email") else "No",
                    lead.get("phone", ""),
                    social_links,
                    lead.get("score", ""),
                    lead.get("priority", ""),
                ])
